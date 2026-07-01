from __future__ import annotations

from io import BytesIO
import json
import os
from pathlib import Path
from datetime import datetime
from typing import Any

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

from excel_export import build_workbook

try:
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    from googleapiclient.http import MediaFileUpload
except ImportError:
    Credentials = None
    build = None
    MediaIoBaseDownload = None
    MediaFileUpload = None


APP_TITLE = "Khảo sát du lịch công ty Meiwa"
DESTINATIONS = ["Nha Trang", "Đà Lạt"]
DEPARTMENTS = ["GA", "CR", "CS", "CD", "PT", "QA", "MOLD"]
EXPORT_FILE_NAME = "Form khao sat Du lich Cong ty meiwa nam 2026.xlsx"
EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def init_state() -> None:
    if "records" not in st.session_state:
        st.session_state.records = []
    if "dirty_export" not in st.session_state:
        st.session_state.dirty_export = False
    if "last_saved_path" not in st.session_state:
        st.session_state.last_saved_path = ""
    if "last_saved_at" not in st.session_state:
        st.session_state.last_saved_at = ""
    if "last_drive_url" not in st.session_state:
        st.session_state.last_drive_url = ""
    if "input_msnv" not in st.session_state:
        st.session_state.input_msnv = ""
    if "input_ho_ten" not in st.session_state:
        st.session_state.input_ho_ten = ""
    if "input_bo_phan" not in st.session_state:
        st.session_state.input_bo_phan = DEPARTMENTS[3]
    if "input_cong_doan" not in st.session_state:
        st.session_state.input_cong_doan = ""
    if "input_tham_gia" not in st.session_state:
        st.session_state.input_tham_gia = None
    if "input_dia_diem" not in st.session_state:
        st.session_state.input_dia_diem = None
    if "reset_form_pending" not in st.session_state:
        st.session_state.reset_form_pending = False


def reset_form_inputs() -> None:
    st.session_state.input_msnv = ""
    st.session_state.input_ho_ten = ""
    st.session_state.input_bo_phan = DEPARTMENTS[3]
    st.session_state.input_cong_doan = ""
    st.session_state.input_tham_gia = None
    st.session_state.input_dia_diem = None


def normalize_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_name(value: object) -> str:
    raw = " ".join(normalize_text(value).split())
    if not raw:
        return ""
    return " ".join(part[:1].upper() + part[1:].lower() if part else "" for part in raw.split(" "))


def normalize_department(value: object) -> str:
    raw = normalize_text(value).upper().replace(".", "").replace(" ", "")
    if not raw:
        return ""
    aliases = {
        "MOLD": "MOLD",
        "GA": "GA",
        "CR": "CR",
        "CS": "CS",
        "CD": "CD",
        "PT": "PT",
        "QA": "QA",
    }
    return aliases.get(raw, raw)


def is_joining(value: object) -> bool:
    return normalize_text(value).lower() in {"co", "có", "yes", "y", "1", "true"}


def record_df(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(
            columns=["MSNV", "Họ tên", "Bộ phận", "Công đoạn", "Tham gia", "Địa điểm"]
        )
    frame = pd.DataFrame(records)
    frame = frame.rename(
        columns={
            "msnv": "MSNV",
            "ho_ten": "Họ tên",
            "bo_phan": "Bộ phận",
            "cong_doan": "Công đoạn",
            "tham_gia": "Tham gia",
            "dia_diem": "Địa điểm",
        }
    )
    wanted = ["MSNV", "Họ tên", "Bộ phận", "Công đoạn", "Tham gia", "Địa điểm"]
    for col in wanted:
        if col not in frame.columns:
            frame[col] = ""
    frame = frame[wanted]
    return frame


def upsert_record(record: dict) -> None:
    records = st.session_state.records
    key = normalize_text(record.get("msnv"))
    if not key:
        return
    replaced = False
    for idx, existing in enumerate(records):
        if normalize_text(existing.get("msnv")) == key:
            records[idx] = record
            replaced = True
            break
    if not replaced:
        records.append(record)


def metrics(records: list[dict]) -> dict[str, int]:
    yes = 0
    no = 0
    nha_trang = 0
    da_lat = 0
    for item in records:
        participate = normalize_text(item.get("tham_gia")).lower() in {"co", "có", "yes", "y", "1", "true"}
        destination = normalize_text(item.get("dia_diem")).lower()
        if participate:
            yes += 1
            if destination == "nha trang":
                nha_trang += 1
            elif destination in {"đà lạt", "da lat"}:
                da_lat += 1
        else:
            no += 1
    return {
        "total": len(records),
        "yes": yes,
        "no": no,
        "nha_trang": nha_trang,
        "da_lat": da_lat,
    }


def department_summary(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=["Bộ phận", "Tổng phiếu", "Tham gia"])

    frame = pd.DataFrame(records)
    if "bo_phan" in frame.columns:
        frame["bo_phan"] = frame["bo_phan"].map(normalize_department)
    else:
        frame["bo_phan"] = ""
    if "tham_gia" in frame.columns:
        frame["tham_gia"] = frame["tham_gia"].astype(str).str.strip().str.lower()
    else:
        frame["tham_gia"] = ""
    frame["is_join"] = frame["tham_gia"].isin(["co", "có", "yes", "y", "1", "true"])

    grouped = (
        frame.groupby("bo_phan", dropna=False)
        .agg(Tổng_phiếu=("msnv", "count"), Tham_gia=("is_join", "sum"))
        .reset_index()
        .rename(columns={"bo_phan": "Bộ phận", "Tổng_phiếu": "Tổng phiếu", "Tham_gia": "Tham gia"})
    )

    if not grouped.empty:
        order_map = {dept: idx for idx, dept in enumerate(DEPARTMENTS)}
        grouped["_order"] = grouped["Bộ phận"].map(lambda x: order_map.get(str(x).strip().upper(), 999))
        grouped = grouped.sort_values(["_order", "Bộ phận"]).drop(columns=["_order"]).reset_index(drop=True)
    return grouped


def destination_summary(records: list[dict]) -> pd.DataFrame:
    stats = metrics(records)
    total_join = stats["yes"]
    rows = [
        {
            "Hạng mục": "Tổng tham gia",
            "Số người": total_join,
            "Tỷ lệ": "100%" if total_join else "0%",
        },
        {
            "Hạng mục": "Nha Trang",
            "Số người": stats["nha_trang"],
            "Tỷ lệ": f"{(stats['nha_trang'] / total_join * 100):.1f}%" if total_join else "0%",
        },
        {
            "Hạng mục": "Đà Lạt",
            "Số người": stats["da_lat"],
            "Tỷ lệ": f"{(stats['da_lat'] / total_join * 100):.1f}%" if total_join else "0%",
        },
    ]
    return pd.DataFrame(rows)


def department_destination_summary(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=["Bộ phận", "Tham gia", "Nha Trang", "Đà Lạt"])

    rows: list[dict] = []
    for department in DEPARTMENTS:
        dept_records = [
            item for item in records if normalize_department(item.get("bo_phan")) == department
        ]
        stats = metrics(dept_records)
        if stats["total"] == 0:
            continue
        rows.append(
            {
                "Bộ phận": department,
                "Tham gia": stats["yes"],
                "Nha Trang": stats["nha_trang"],
                "Đà Lạt": stats["da_lat"],
            }
        )
    return pd.DataFrame(rows)


def import_rows_from_frame(frame: pd.DataFrame) -> list[dict]:
    if frame.empty:
        return []

    lookup = {str(col).strip().lower().replace(" ", "").replace("_", ""): col for col in frame.columns}

    def pick(*options: str):
        for option in options:
            key = option.strip().lower().replace(" ", "").replace("_", "")
            if key in lookup:
                return lookup[key]
        return None

    msnv_col = pick("msnv", "ma nv", "manv", "employee id")
    name_col = pick("họ tên", "ho ten", "hoten", "name")
    dept_col = pick("bộ phận", "bo phan", "bophan", "department")
    dept_process_col = pick("công đoạn", "cong doan", "process", "stage")
    join_col = pick("tham gia", "thamgia", "join", "participate")
    dest_col = pick("địa điểm", "dia diem", "destination")

    rows: list[dict] = []
    for _, row in frame.iterrows():
        msnv = normalize_text(row[msnv_col]) if msnv_col else ""
        name = normalize_name(row[name_col]) if name_col else ""
        if not (msnv or name):
            continue
        join_raw = normalize_text(row[join_col]) if join_col else ""
        join_raw_lower = join_raw.lower()
        if (not join_raw) and (normalize_text(row[dest_col]) if dest_col else ""):
            join_raw = "Có"
            join_raw_lower = "có"
        participate = join_raw_lower in {"co", "có", "yes", "y", "1", "true", "x", "v"}
        destination = normalize_text(row[dest_col]) if dest_col else ""
        if not participate:
            destination = ""
            people = 0
        rows.append(
            {
                "msnv": msnv,
                "ho_ten": normalize_name(name),
                "bo_phan": normalize_department(row[dept_col]) if dept_col else "",
                "cong_doan": normalize_text(row[dept_process_col]) if dept_process_col else "",
                "tham_gia": "Có" if participate else "Không",
                "dia_diem": destination if participate else "",
            }
        )
    return rows


def build_export_bytes(records: list[dict]) -> bytes:
    workbook = build_workbook(records)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def get_export_path() -> Path:
    data_dir = os.getenv("DATA_DIR", "")
    export_dir = Path(data_dir) if data_dir else Path(__file__).with_name("exports")
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir / EXPORT_FILE_NAME


def load_existing_records() -> list[dict]:
    export_path = get_export_path()
    if not export_path.exists():
        return []

    workbook = load_workbook(export_path, data_only=False)
    sheet = workbook.active
    records: list[dict] = []

    for row_num in range(9, sheet.max_row + 1):
        msnv = normalize_text(sheet.cell(row_num, 2).value)
        if msnv.upper() == "TOTAL":
            break

        ho_ten = normalize_name(sheet.cell(row_num, 3).value)
        bo_phan = normalize_department(sheet.cell(row_num, 4).value)
        cong_doan = normalize_text(sheet.cell(row_num, 5).value)
        di_co = normalize_text(sheet.cell(row_num, 6).value).lower() == "v"
        di_khong = normalize_text(sheet.cell(row_num, 7).value).lower() == "v"
        nha_trang = normalize_text(sheet.cell(row_num, 8).value).lower() == "v"
        da_lat = normalize_text(sheet.cell(row_num, 9).value).lower() == "v"

        if not any([msnv, ho_ten, bo_phan, cong_doan, di_co, di_khong, nha_trang, da_lat]):
            continue

        tham_gia = "Có" if di_co else "Không"
        dia_diem = "Nha Trang" if nha_trang else "Đà Lạt" if da_lat else ""
        records.append(
            {
                "msnv": msnv,
                "ho_ten": ho_ten,
                "bo_phan": bo_phan,
                "cong_doan": cong_doan,
                "tham_gia": tham_gia,
                "dia_diem": dia_diem if tham_gia == "Có" else "",
            }
        )

    return records


def cleanup_old_export_files(export_dir: Path, keep_name: str) -> None:
    for file_path in export_dir.glob("*.xlsx"):
        if file_path.name != keep_name:
            try:
                file_path.unlink()
            except OSError:
                continue


def read_google_config_from_env() -> tuple[dict[str, Any], dict[str, Any]]:
    drive_settings = {
        "folder_id": os.getenv("GOOGLE_DRIVE_FOLDER_ID", ""),
        "shared_url": os.getenv("GOOGLE_DRIVE_SHARED_URL", ""),
    }
    service_account_raw = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    service_account: dict[str, Any] = {}
    if service_account_raw:
        try:
            service_account = json.loads(service_account_raw)
        except json.JSONDecodeError:
            service_account = {}
    return drive_settings, service_account


def get_google_drive_settings() -> dict[str, str]:
    drive_settings: Any = {}
    service_account: Any = {}

    try:
        if "google_drive" in st.secrets:
            drive_settings = st.secrets["google_drive"]
        if "google_service_account" in st.secrets:
            service_account = st.secrets["google_service_account"]
    except Exception:
        drive_settings, service_account = read_google_config_from_env()
    else:
        if not drive_settings and not service_account:
            drive_settings, service_account = read_google_config_from_env()

    folder_id = normalize_text(drive_settings.get("folder_id", ""))
    shared_url = normalize_text(drive_settings.get("shared_url", ""))
    enabled = bool(folder_id and service_account)

    return {
        "folder_id": folder_id,
        "shared_url": shared_url,
        "enabled": "1" if enabled else "0",
    }


@st.cache_resource(show_spinner=False)
def get_google_drive_service():
    if Credentials is None or build is None:
        return None
    service_account_info: dict[str, Any] | None = None
    try:
        has_service_account = "google_service_account" in st.secrets
    except Exception:
        has_service_account = False
    if has_service_account:
        service_account_info = dict(st.secrets["google_service_account"])
    else:
        _, env_service_account = read_google_config_from_env()
        if env_service_account:
            service_account_info = env_service_account
    if not service_account_info:
        return None
    credentials = Credentials.from_service_account_info(service_account_info, scopes=["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=credentials)


def sync_export_to_google_drive(local_path: Path) -> dict[str, str]:
    settings = get_google_drive_settings()
    if settings["enabled"] != "1":
        return {"status": "not_configured", "url": settings["shared_url"]}

    service = get_google_drive_service()
    if service is None or MediaFileUpload is None:
        return {"status": "missing_library", "url": settings["shared_url"]}

    file_name = local_path.name.replace("'", "\\'")
    query = (
        f"name = '{file_name}' and "
        f"'{settings['folder_id']}' in parents and trashed = false"
    )
    existing = (
        service.files()
        .list(
            q=query,
            spaces="drive",
            fields="files(id, webViewLink, webContentLink)",
            pageSize=1,
        )
        .execute()
        .get("files", [])
    )
    media = MediaFileUpload(str(local_path), mimetype=EXCEL_MIME_TYPE, resumable=False)

    if existing:
        file_id = existing[0]["id"]
        updated = (
            service.files()
            .update(fileId=file_id, media_body=media, fields="id, webViewLink, webContentLink")
            .execute()
        )
        url = updated.get("webViewLink") or updated.get("webContentLink") or settings["shared_url"]
        return {"status": "updated", "url": url}

    created = (
        service.files()
        .create(
            body={"name": local_path.name, "parents": [settings["folder_id"]]},
            media_body=media,
            fields="id, webViewLink, webContentLink",
        )
        .execute()
    )
    url = created.get("webViewLink") or created.get("webContentLink") or settings["shared_url"]
    return {"status": "created", "url": url}


def ensure_local_export_from_google_drive() -> dict[str, str]:
    export_path = get_export_path()
    if export_path.exists():
        return {"status": "local_exists", "path": str(export_path)}

    settings = get_google_drive_settings()
    if settings["enabled"] != "1":
        return {"status": "not_configured", "path": str(export_path)}

    service = get_google_drive_service()
    if service is None or MediaIoBaseDownload is None:
        return {"status": "missing_library", "path": str(export_path)}

    file_name = export_path.name.replace("'", "\\'")
    query = (
        f"name = '{file_name}' and "
        f"'{settings['folder_id']}' in parents and trashed = false"
    )
    existing = (
        service.files()
        .list(
            q=query,
            spaces="drive",
            fields="files(id, name)",
            pageSize=1,
        )
        .execute()
        .get("files", [])
    )
    if not existing:
        return {"status": "drive_file_missing", "path": str(export_path)}

    request = service.files().get_media(fileId=existing[0]["id"])
    export_path.parent.mkdir(parents=True, exist_ok=True)
    with export_path.open("wb") as file_obj:
        downloader = MediaIoBaseDownload(file_obj, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    return {"status": "downloaded", "path": str(export_path)}


def save_export_file(records: list[dict]) -> Path:
    export_dir = get_export_path().parent
    cleanup_old_export_files(export_dir, EXPORT_FILE_NAME)
    export_path = get_export_path()
    workbook = build_workbook(records)
    workbook.save(export_path)
    return export_path


def render_close_warning(is_dirty: bool) -> None:
    components.html(
        f"""
        <script>
        window.onbeforeunload = {str(is_dirty).lower()}
          ? function(event) {{
              event.preventDefault();
              event.returnValue = "";
              return "";
            }}
          : null;
        </script>
        """,
        height=0,
    )


def load_rows_from_upload(uploaded_file) -> list[dict]:
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix == ".csv":
        frame = pd.read_csv(uploaded_file)
    else:
        frame = pd.read_excel(uploaded_file)
    return import_rows_from_frame(frame)


def department_summary(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=["B盻・ph蘯ｭn", "T盻貧g phi蘯ｿu", "Tham gia", "Khﾃｴng tham gia"])

    frame = pd.DataFrame(records)
    frame["bo_phan"] = frame.get("bo_phan", "").map(normalize_department) if "bo_phan" in frame.columns else ""
    frame["tham_gia"] = frame.get("tham_gia", "").astype(str).str.strip().str.lower() if "tham_gia" in frame.columns else ""
    frame["is_join"] = frame["tham_gia"].isin(["co", "cﾃｳ", "yes", "y", "1", "true"])

    grouped = (
        frame.groupby("bo_phan", dropna=False)
        .agg(T盻貧g_phi蘯ｿu=("msnv", "count"), Tham_gia=("is_join", "sum"))
        .reset_index()
        .rename(columns={"bo_phan": "B盻・ph蘯ｭn", "T盻貧g_phi蘯ｿu": "T盻貧g phi蘯ｿu", "Tham_gia": "Tham gia"})
    )
    grouped["Khﾃｴng tham gia"] = grouped["T盻貧g phi蘯ｿu"] - grouped["Tham gia"]

    order_map = {dept: idx for idx, dept in enumerate(DEPARTMENTS)}
    grouped["_order"] = grouped["B盻・ph蘯ｭn"].map(lambda x: order_map.get(str(x).strip().upper(), 999))
    return grouped.sort_values(["_order", "B盻・ph蘯ｭn"]).drop(columns=["_order"]).reset_index(drop=True)


def destination_summary(records: list[dict]) -> pd.DataFrame:
    stats = metrics(records)
    total_join = stats["yes"]
    rows = [
        {
            "H蘯｡ng m盻･c": "T盻貧g tham gia",
            "S盻・ngﾆｰ盻拱": total_join,
            "T盻ｷ l盻・": "100%" if total_join else "0%",
        },
        {
            "H蘯｡ng m盻･c": "Khﾃｴng tham gia",
            "S盻・ngﾆｰ盻拱": stats["no"],
            "T盻ｷ l盻・": f"{(stats['no'] / stats['total'] * 100):.1f}%" if stats["total"] else "0%",
        },
        {
            "H蘯｡ng m盻･c": "Nha Trang",
            "S盻・ngﾆｰ盻拱": stats["nha_trang"],
            "T盻ｷ l盻・": f"{(stats['nha_trang'] / total_join * 100):.1f}%" if total_join else "0%",
        },
        {
            "H蘯｡ng m盻･c": "ﾄ静 L蘯｡t",
            "S盻・ngﾆｰ盻拱": stats["da_lat"],
            "T盻ｷ l盻・": f"{(stats['da_lat'] / total_join * 100):.1f}%" if total_join else "0%",
        },
    ]
    return pd.DataFrame(rows)


def department_destination_summary(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=["B盻・ph蘯ｭn", "Tham gia", "Khﾃｴng tham gia", "Nha Trang", "ﾄ静 L蘯｡t"])

    rows: list[dict] = []
    for department in DEPARTMENTS:
        dept_records = [
            item for item in records if normalize_department(item.get("bo_phan")) == department
        ]
        stats = metrics(dept_records)
        if stats["total"] == 0:
            continue
        rows.append(
            {
                "B盻・ph蘯ｭn": department,
                "Tham gia": stats["yes"],
                "Khﾃｴng tham gia": stats["no"],
                "Nha Trang": stats["nha_trang"],
                "ﾄ静 L蘯｡t": stats["da_lat"],
            }
        )
    return pd.DataFrame(rows)


def render_shell() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="🧳", layout="centered")
    st.markdown(
        """
        <style>
        .block-container {
            padding-top: 1.25rem;
            padding-bottom: 2rem;
            max-width: 760px;
        }
        .hero {
            border-radius: 24px;
            padding: 1.2rem 1.4rem;
            background: linear-gradient(135deg, #fff6ea 0%, #eef6ff 48%, #f6ecff 100%);
            border: 1px solid rgba(17, 24, 39, 0.08);
            box-shadow: 0 16px 40px rgba(15, 23, 42, 0.08);
            margin-bottom: 1rem;
        }
        .hero h1 {
            margin: 0;
            font-size: 2rem;
            line-height: 1.15;
        }
        .hero p {
            margin: 0.35rem 0 0;
            color: #475569;
        }
        .card {
            background: #ffffff;
            border: 1px solid rgba(15, 23, 42, 0.08);
            border-radius: 20px;
            padding: 1rem 1rem 0.85rem;
            box-shadow: 0 10px 28px rgba(15, 23, 42, 0.06);
            margin-bottom: 1rem;
        }
        .save-alert {
            border: 2px solid #dc2626;
            background: linear-gradient(135deg, #fff7ed 0%, #fef2f2 100%);
        }
        .save-title {
            color: #b91c1c;
            font-size: 1.1rem;
            font-weight: 800;
            margin-bottom: 0.35rem;
        }
        .save-copy {
            color: #7f1d1d;
            font-size: 0.95rem;
            margin-bottom: 0.8rem;
        }
        .section-title {
            font-size: 1rem;
            font-weight: 700;
            margin: 0 0 0.6rem;
        }
        .hint {
            color: #64748b;
            font-size: 0.92rem;
            margin-top: 0.25rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    init_state()
    if st.session_state.reset_form_pending:
        reset_form_inputs()
        st.session_state.reset_form_pending = False
    if "bootstrap_drive_status" not in st.session_state:
        st.session_state.bootstrap_drive_status = ensure_local_export_from_google_drive().get("status", "")
    st.session_state.records = load_existing_records()
    render_close_warning(st.session_state.dirty_export)

    st.markdown(
        f"""
        <div class="hero">
          <h1>{APP_TITLE}</h1>
          <p>Mở app là nhập ngay. Nếu chọn Không, phần địa điểm tự ẩn.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="card">', unsafe_allow_html=True)
    msnv = st.text_input("MSNV", placeholder="Ví dụ: G06071209", key="input_msnv")
    ho_ten = st.text_input("Họ tên", placeholder="Ví dụ: Lê Hữu Phước", key="input_ho_ten")
    bo_phan = st.selectbox("Bộ phận", DEPARTMENTS, key="input_bo_phan")
    cong_doan = st.text_input("Công đoạn", placeholder="Ví dụ: 000 hoặc 3-5", key="input_cong_doan")
    tham_gia = st.selectbox(
        "Bạn có tham gia chuyến đi không?",
        ["Có", "Không"],
        index=None,
        placeholder="Chọn Có hoặc Không",
        key="input_tham_gia",
    )
    dia_diem = ""
    if is_joining(tham_gia):
        dia_diem = st.selectbox(
            "Chọn 1 địa điểm",
            DESTINATIONS,
            index=None,
            placeholder="Chọn Nha Trang hoặc Đà Lạt",
            key="input_dia_diem",
        )

    submitted = st.button("LƯU KẾT QUẢ KHẢO SÁT", width="stretch")
    if submitted:
        if not msnv or not ho_ten:
            st.error("Nhập ít nhất MSNV và Họ tên.")
        elif not tham_gia:
            st.error("Chọn Có hoặc Không tham gia.")
        elif is_joining(tham_gia) and not dia_diem:
            st.error("Chọn địa điểm khi người đó tham gia.")
        else:
            record = {
                "msnv": msnv,
                "ho_ten": normalize_name(ho_ten),
                "bo_phan": normalize_department(bo_phan),
                "cong_doan": cong_doan,
                "tham_gia": "Có" if is_joining(tham_gia) else "Không",
                "dia_diem": dia_diem if is_joining(tham_gia) else "",
            }
            st.session_state.records = load_existing_records()
            upsert_record(record)
            export_path = save_export_file(st.session_state.records)
            drive_result = {"status": "not_configured", "url": ""}
            try:
                drive_result = sync_export_to_google_drive(export_path)
            except Exception as exc:
                drive_result = {"status": f"error: {exc}", "url": ""}
            st.session_state.last_saved_path = str(export_path)
            st.session_state.last_saved_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            st.session_state.last_drive_url = drive_result.get("url", "")
            st.session_state.dirty_export = False
            st.session_state.reset_form_pending = True
            if drive_result["status"] in {"created", "updated"}:
                st.success(
                    f"Đã lưu khảo sát, cập nhật file tổng và đồng bộ Google Drive lúc {st.session_state.last_saved_at}"
                )
            elif drive_result["status"] == "not_configured":
                st.success(
                    f"Đã lưu khảo sát và cập nhật file tổng lúc {st.session_state.last_saved_at}"
                )
            elif drive_result["status"] == "missing_library":
                st.warning("Đã lưu file tổng, nhưng máy chủ chưa cài thư viện Google Drive để đồng bộ.")
            else:
                st.warning(
                    f"Đã lưu file tổng, nhưng chưa đẩy được lên Google Drive: {drive_result['status']}"
                )
            st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card save-alert">', unsafe_allow_html=True)
    st.markdown('<div class="save-title">TỰ ĐỘNG LƯU FILE TỔNG</div>', unsafe_allow_html=True)
    st.markdown('<div class="save-copy">Mỗi lần bấm "LƯU KẾT QUẢ KHẢO SÁT", app sẽ tự cập nhật ngay vào file Excel tổng.</div>', unsafe_allow_html=True)
    excel_bytes = build_export_bytes(st.session_state.records)
    st.download_button(
        "TẢI FILE TỔNG",
        data=excel_bytes,
        file_name=EXPORT_FILE_NAME,
        mime=EXCEL_MIME_TYPE,
        width="stretch",
    )
    if st.session_state.last_saved_path:
        st.caption(f"File tổng mới nhất: `{st.session_state.last_saved_path}`")
        st.markdown(f"[Mở file tổng đã lưu]({Path(st.session_state.last_saved_path).as_uri()})")
    drive_settings = get_google_drive_settings()
    if drive_settings["enabled"] == "1":
        st.caption("Google Drive: đã cấu hình đồng bộ tự động.")
    else:
        st.caption("Google Drive: chưa cấu hình. App hiện vẫn lưu file tổng trong máy và cho tải trực tiếp.")
    drive_url = st.session_state.last_drive_url or drive_settings["shared_url"]
    if drive_url:
        st.markdown(f"[MỞ FILE TỔNG TRÊN GOOGLE DRIVE]({drive_url})")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Đã nhập</div>', unsafe_allow_html=True)
    if st.session_state.records:
        rows = [
            {
                "MSNV": item.get("msnv", ""),
                "Họ tên": item.get("ho_ten", ""),
                "Bộ phận": item.get("bo_phan", ""),
                "Công đoạn": item.get("cong_doan", ""),
                "Đi": item.get("tham_gia", ""),
                "Địa điểm": item.get("dia_diem", ""),
            }
            for item in st.session_state.records
        ]
        st.dataframe(rows, width="stretch", hide_index=True)
    else:
        st.info("Chưa có dữ liệu.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Biểu đồ tổng toàn công ty</div>', unsafe_allow_html=True)
    stats = metrics(st.session_state.records)
    overview = destination_summary(st.session_state.records)
    if stats["yes"] > 0:
        col1, col2, col3 = st.columns(3)
        col1.metric("Tổng người tham gia", stats["yes"])
        col2.metric("Nha Trang", stats["nha_trang"], f"{(stats['nha_trang'] / stats['yes'] * 100):.1f}%")
        col3.metric("Đà Lạt", stats["da_lat"], f"{(stats['da_lat'] / stats['yes'] * 100):.1f}%")
        st.dataframe(overview, width="stretch", hide_index=True)
        chart_frame = overview[overview["Hạng mục"] != "Tổng tham gia"].set_index("Hạng mục")[["Số người"]]
        st.bar_chart(chart_frame, width="stretch")
    else:
        st.info("Chưa có người chọn tham gia để hiển thị biểu đồ địa điểm.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Biểu đồ từng bộ phận</div>', unsafe_allow_html=True)
    dept_frame = department_summary(st.session_state.records)
    if not dept_frame.empty:
        st.dataframe(dept_frame, width="stretch", hide_index=True)
        chart_frame = dept_frame.set_index("Bộ phận")[["Tham gia"]]
        st.bar_chart(chart_frame, width="stretch")
        destination_by_department = department_destination_summary(st.session_state.records)
        if not destination_by_department.empty:
            st.dataframe(destination_by_department, width="stretch", hide_index=True)
            detail_chart_frame = destination_by_department.set_index("Bộ phận")[["Nha Trang", "Đà Lạt"]]
            st.bar_chart(detail_chart_frame, width="stretch")
    else:
        st.info("Chưa có dữ liệu để hiển thị dashboard bộ phận.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Bi盻ブ b盻・sung: tham gia vﾃ khﾃｴng tham gia</div>', unsafe_allow_html=True)
    company_compare = pd.DataFrame(
        [
            {"H蘯｡ng m盻･c": "Tham gia", "S盻・ngﾆｰ盻拱": stats["yes"]},
            {"H蘯｡ng m盻･c": "Khﾃｴng tham gia", "S盻・ngﾆｰ盻拱": stats["no"]},
            {"H蘯｡ng m盻･c": "Nha Trang", "S盻・ngﾆｰ盻拱": stats["nha_trang"]},
            {"H蘯｡ng m盻･c": "ﾄ静 L蘯｡t", "S盻・ngﾆｰ盻拱": stats["da_lat"]},
        ]
    )
    st.dataframe(company_compare, width="stretch", hide_index=True)
    st.bar_chart(company_compare.set_index("H蘯｡ng m盻･c")[["S盻・ngﾆｰ盻拱"]], width="stretch")

    destination_by_department = department_destination_summary(st.session_state.records)
    if not destination_by_department.empty:
        st.markdown('<div class="hint">Chi ti盻ｿt theo b盻・ph蘯ｭn: hi盻ハ th盻・c蘯｣ tham gia, khﾃｴng tham gia vﾃ ﾄ黛ｻ蟻 ﾄ訴盻ノ ﾄ妥｣ ch盻肱.</div>', unsafe_allow_html=True)
        st.dataframe(destination_by_department, width="stretch", hide_index=True)
        st.bar_chart(
            destination_by_department.set_index("B盻・ph蘯ｭn")[["Tham gia", "Khﾃｴng tham gia", "Nha Trang", "ﾄ静 L蘯｡t"]],
            width="stretch",
        )
    st.markdown("</div>", unsafe_allow_html=True)


    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Dashboard tong hop de doc</div>', unsafe_allow_html=True)
    clean_company = pd.DataFrame(
        [
            {"Hang muc": "Tong phieu", "So nguoi": stats["total"]},
            {"Hang muc": "Tong tham gia", "So nguoi": stats["yes"]},
            {"Hang muc": "Khong tham gia", "So nguoi": stats["no"]},
            {"Hang muc": "Nha Trang", "So nguoi": stats["nha_trang"]},
            {"Hang muc": "Da Lat", "So nguoi": stats["da_lat"]},
        ]
    )
    st.caption("Bang tong hop toan cong ty de doi chieu nhanh.")
    st.dataframe(clean_company, width="stretch", hide_index=True)
    st.bar_chart(clean_company.set_index("Hang muc")[["So nguoi"]], width="stretch")

    clean_department_rows = []
    for dept in DEPARTMENTS:
        dept_items = [
            item for item in st.session_state.records if normalize_department(item.get("bo_phan")) == dept
        ]
        dept_stats = metrics(dept_items)
        if dept_stats["total"] == 0:
            continue
        clean_department_rows.append(
            {
                "Bo phan": dept,
                "Tong phieu": dept_stats["total"],
                "Tham gia": dept_stats["yes"],
                "Khong tham gia": dept_stats["no"],
                "Nha Trang": dept_stats["nha_trang"],
                "Da Lat": dept_stats["da_lat"],
            }
        )

    clean_department = pd.DataFrame(clean_department_rows)
    if not clean_department.empty:
        total_detail = pd.DataFrame(
            [
                {
                    "Bo phan": "TONG",
                    "Tong phieu": int(clean_department["Tong phieu"].sum()),
                    "Tham gia": int(clean_department["Tham gia"].sum()),
                    "Khong tham gia": int(clean_department["Khong tham gia"].sum()),
                    "Nha Trang": int(clean_department["Nha Trang"].sum()),
                    "Da Lat": int(clean_department["Da Lat"].sum()),
                }
            ]
        )
        clean_department_view = pd.concat([clean_department, total_detail], ignore_index=True)
        st.markdown('<div class="section-title">Chi tiet tung bo phan co dong TONG</div>', unsafe_allow_html=True)
        st.caption("Dong TONG ben duoi phai khop voi so tong quat ben tren.")
        st.dataframe(clean_department_view, width="stretch", hide_index=True)
        st.bar_chart(
            clean_department.set_index("Bo phan")[["Tong phieu", "Tham gia", "Khong tham gia", "Nha Trang", "Da Lat"]],
            width="stretch",
        )
    st.markdown("</div>", unsafe_allow_html=True)


def _normalize_destination_label(value: object) -> str:
    raw = normalize_text(value).lower()
    if raw == "nha trang":
        return "Nha Trang"
    if raw in {"đà lạt", "da lat", "ﾄ妥 l蘯｡t", "・・撕・ｰ l陂ｯ・｡t"}:
        return "Đà Lạt"
    return ""


def _normalize_join_label(value: object) -> str:
    return "Có" if normalize_text(value).lower() in {"co", "có", "cﾃｳ", "yes", "y", "1", "true", "x", "v"} else "Không"


def _clean_records(records: list[dict]) -> list[dict]:
    cleaned: list[dict] = []
    for item in records:
        join_label = _normalize_join_label(item.get("tham_gia"))
        cleaned.append(
            {
                "msnv": normalize_text(item.get("msnv")),
                "ho_ten": normalize_name(item.get("ho_ten")),
                "bo_phan": normalize_department(item.get("bo_phan")),
                "cong_doan": normalize_text(item.get("cong_doan")),
                "tham_gia": join_label,
                "dia_diem": _normalize_destination_label(item.get("dia_diem")) if join_label == "Có" else "",
            }
        )
    return cleaned


def _dashboard_metrics(records: list[dict]) -> dict[str, int]:
    cleaned = _clean_records(records)
    yes = sum(1 for item in cleaned if item["tham_gia"] == "Có")
    no = sum(1 for item in cleaned if item["tham_gia"] == "Không")
    nha_trang = sum(1 for item in cleaned if item["dia_diem"] == "Nha Trang")
    da_lat = sum(1 for item in cleaned if item["dia_diem"] == "Đà Lạt")
    return {
        "total": len(cleaned),
        "yes": yes,
        "no": no,
        "nha_trang": nha_trang,
        "da_lat": da_lat,
    }


def _company_summary_frame(records: list[dict]) -> pd.DataFrame:
    stats = _dashboard_metrics(records)
    total = stats["total"]
    yes = stats["yes"]
    return pd.DataFrame(
        [
            {
                "Hạng mục": "Tổng phiếu",
                "Số lượng": stats["total"],
                "Tỷ lệ": "100%" if total else "0%",
            },
            {
                "Hạng mục": "Tổng tham gia",
                "Số lượng": stats["yes"],
                "Tỷ lệ": f"{(stats['yes'] / total * 100):.1f}%" if total else "0%",
            },
            {
                "Hạng mục": "Không tham gia",
                "Số lượng": stats["no"],
                "Tỷ lệ": f"{(stats['no'] / total * 100):.1f}%" if total else "0%",
            },
            {
                "Hạng mục": "Nha Trang",
                "Số lượng": stats["nha_trang"],
                "Tỷ lệ": f"{(stats['nha_trang'] / yes * 100):.1f}%" if yes else "0%",
            },
            {
                "Hạng mục": "Đà Lạt",
                "Số lượng": stats["da_lat"],
                "Tỷ lệ": f"{(stats['da_lat'] / yes * 100):.1f}%" if yes else "0%",
            },
        ]
    )


def _department_summary_frame(records: list[dict]) -> pd.DataFrame:
    cleaned = _clean_records(records)
    rows: list[dict[str, int | str]] = []
    for department in DEPARTMENTS:
        dept_records = [item for item in cleaned if item["bo_phan"] == department]
        if not dept_records:
            continue
        stats = _dashboard_metrics(dept_records)
        rows.append(
            {
                "Bộ phận": department,
                "Tổng phiếu": stats["total"],
                "Tham gia": stats["yes"],
                "Không tham gia": stats["no"],
                "Nha Trang": stats["nha_trang"],
                "Đà Lạt": stats["da_lat"],
            }
        )

    if not rows:
        return pd.DataFrame(columns=["Bộ phận", "Tổng phiếu", "Tham gia", "Không tham gia", "Nha Trang", "Đà Lạt"])

    frame = pd.DataFrame(rows)
    total_row = pd.DataFrame(
        [
            {
                "Bộ phận": "TỔNG",
                "Tổng phiếu": int(frame["Tổng phiếu"].sum()),
                "Tham gia": int(frame["Tham gia"].sum()),
                "Không tham gia": int(frame["Không tham gia"].sum()),
                "Nha Trang": int(frame["Nha Trang"].sum()),
                "Đà Lạt": int(frame["Đà Lạt"].sum()),
            }
        ]
    )
    return pd.concat([frame, total_row], ignore_index=True)


def render_shell() -> None:
    st.set_page_config(page_title="Khảo sát du lịch công ty Meiwa", page_icon="🧳", layout="centered")
    st.markdown(
        """
        <style>
        .block-container {
            padding-top: 1.25rem;
            padding-bottom: 2rem;
            max-width: 840px;
        }
        .hero {
            border-radius: 24px;
            padding: 1.2rem 1.4rem;
            background: linear-gradient(135deg, #fff6ea 0%, #eef6ff 48%, #f6ecff 100%);
            border: 1px solid rgba(17, 24, 39, 0.08);
            box-shadow: 0 16px 40px rgba(15, 23, 42, 0.08);
            margin-bottom: 1rem;
        }
        .hero h1 {
            margin: 0;
            font-size: 2rem;
            line-height: 1.15;
        }
        .hero p {
            margin: 0.35rem 0 0;
            color: #475569;
        }
        .card {
            background: #ffffff;
            border: 1px solid rgba(15, 23, 42, 0.08);
            border-radius: 20px;
            padding: 1rem 1rem 0.85rem;
            box-shadow: 0 10px 28px rgba(15, 23, 42, 0.06);
            margin-bottom: 1rem;
        }
        .save-alert {
            border: 2px solid #dc2626;
            background: linear-gradient(135deg, #fff7ed 0%, #fef2f2 100%);
        }
        .save-title {
            color: #b91c1c;
            font-size: 1.1rem;
            font-weight: 800;
            margin-bottom: 0.35rem;
        }
        .save-copy {
            color: #7f1d1d;
            font-size: 0.95rem;
            margin-bottom: 0.8rem;
        }
        .section-title {
            font-size: 1.08rem;
            font-weight: 800;
            margin: 0 0 0.6rem;
        }
        .hint {
            color: #64748b;
            font-size: 0.92rem;
            margin-top: 0.25rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    APP_TITLE = "Khảo sát du lịch công ty Meiwa"
    DESTINATIONS = ["Nha Trang", "Đà Lạt"]

    init_state()
    if st.session_state.reset_form_pending:
        reset_form_inputs()
        st.session_state.reset_form_pending = False
    if "bootstrap_drive_status" not in st.session_state:
        st.session_state.bootstrap_drive_status = ensure_local_export_from_google_drive().get("status", "")

    st.session_state.records = _clean_records(load_existing_records())
    render_close_warning(st.session_state.dirty_export)

    def mark_dirty() -> None:
        st.session_state.dirty_export = True

    st.markdown(
        f"""
        <div class="hero">
          <h1>{APP_TITLE}</h1>
          <p>Mở app là nhập ngay. Nếu chọn Không, phần địa điểm sẽ tự ẩn và không cần chọn nơi đi.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="card">', unsafe_allow_html=True)
    msnv = st.text_input("MSNV", placeholder="Ví dụ: G06071209", key="input_msnv", on_change=mark_dirty)
    ho_ten = st.text_input("Họ tên", placeholder="Ví dụ: Nguyễn Duy Hoà", key="input_ho_ten", on_change=mark_dirty)
    bo_phan = st.selectbox("Bộ phận", DEPARTMENTS, key="input_bo_phan", on_change=mark_dirty)
    cong_doan = st.text_input("Công đoạn", placeholder="Ví dụ: 56-0 hoặc 3-5", key="input_cong_doan", on_change=mark_dirty)
    tham_gia = st.selectbox(
        "Bạn có tham gia chuyến đi không?",
        ["Có", "Không"],
        index=None,
        placeholder="Chọn Có hoặc Không",
        key="input_tham_gia",
        on_change=mark_dirty,
    )

    dia_diem = ""
    if tham_gia == "Có":
        dia_diem = st.selectbox(
            "Chọn 1 địa điểm",
            DESTINATIONS,
            index=None,
            placeholder="Chọn Nha Trang hoặc Đà Lạt",
            key="input_dia_diem",
            on_change=mark_dirty,
        )
    else:
        st.session_state.input_dia_diem = None

    submitted = st.button("LƯU KẾT QUẢ KHẢO SÁT", width="stretch")
    if submitted:
        if not msnv or not ho_ten:
            st.error("Cần nhập MSNV và Họ tên.")
        elif not tham_gia:
            st.error("Cần chọn Có hoặc Không tham gia.")
        elif tham_gia == "Có" and not dia_diem:
            st.error("Người tham gia cần chọn địa điểm.")
        else:
            record = {
                "msnv": normalize_text(msnv),
                "ho_ten": normalize_name(ho_ten),
                "bo_phan": normalize_department(bo_phan),
                "cong_doan": normalize_text(cong_doan),
                "tham_gia": tham_gia,
                "dia_diem": dia_diem if tham_gia == "Có" else "",
            }
            current_records = _clean_records(load_existing_records())
            st.session_state.records = current_records
            upsert_record(record)
            st.session_state.records = _clean_records(st.session_state.records)
            export_path = save_export_file(st.session_state.records)
            drive_result = {"status": "not_configured", "url": ""}
            try:
                drive_result = sync_export_to_google_drive(export_path)
            except Exception as exc:
                drive_result = {"status": f"error: {exc}", "url": ""}
            st.session_state.last_saved_path = str(export_path)
            st.session_state.last_saved_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            st.session_state.last_drive_url = drive_result.get("url", "")
            st.session_state.dirty_export = False
            st.session_state.reset_form_pending = True
            if drive_result["status"] in {"created", "updated"}:
                st.success(f"Đã lưu khảo sát và cập nhật file tổng lúc {st.session_state.last_saved_at}.")
            elif drive_result["status"] == "not_configured":
                st.success(f"Đã lưu khảo sát vào file tổng lúc {st.session_state.last_saved_at}.")
            else:
                st.warning(f"Đã lưu file tổng, nhưng Google Drive chưa cập nhật: {drive_result['status']}")
            st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card save-alert">', unsafe_allow_html=True)
    st.markdown('<div class="save-title">TỰ ĐỘNG LƯU FILE TỔNG</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="save-copy">Mỗi lần bấm "LƯU KẾT QUẢ KHẢO SÁT", app sẽ cập nhật ngay vào 1 file tổng duy nhất.</div>',
        unsafe_allow_html=True,
    )
    excel_bytes = build_export_bytes(st.session_state.records)
    st.download_button(
        "TẢI FILE TỔNG",
        data=excel_bytes,
        file_name=EXPORT_FILE_NAME,
        mime=EXCEL_MIME_TYPE,
        width="stretch",
    )
    drive_settings = get_google_drive_settings()
    drive_url = st.session_state.last_drive_url or drive_settings["shared_url"]
    if st.session_state.last_saved_path:
        st.caption(f"File tổng mới nhất: `{st.session_state.last_saved_path}`")
    if drive_url:
        st.markdown(f"[MỞ FILE TỔNG TRÊN GOOGLE DRIVE]({drive_url})")
    else:
        st.caption("Chưa cấu hình Google Drive. App vẫn lưu file tổng trong máy và tải trực tiếp được.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Danh sách đã nhập</div>', unsafe_allow_html=True)
    records_frame = pd.DataFrame(_clean_records(st.session_state.records))
    if not records_frame.empty:
        display_frame = records_frame.rename(
            columns={
                "msnv": "MSNV",
                "ho_ten": "Họ tên",
                "bo_phan": "Bộ phận",
                "cong_doan": "Công đoạn",
                "tham_gia": "Đi",
                "dia_diem": "Địa điểm",
            }
        )[["MSNV", "Họ tên", "Bộ phận", "Công đoạn", "Đi", "Địa điểm"]]
        st.dataframe(display_frame, width="stretch", hide_index=True)
    else:
        st.info("Chưa có dữ liệu khảo sát.")
    st.markdown("</div>", unsafe_allow_html=True)

    stats = _dashboard_metrics(st.session_state.records)
    company_frame = _company_summary_frame(st.session_state.records)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Biểu đồ tổng quan toàn công ty</div>', unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Tổng phiếu", stats["total"])
    col2.metric("Tham gia", stats["yes"], f"{(stats['yes'] / stats['total'] * 100):.1f}%" if stats["total"] else "0%")
    col3.metric("Không tham gia", stats["no"], f"{(stats['no'] / stats['total'] * 100):.1f}%" if stats["total"] else "0%")
    col4.metric("Địa điểm đã chọn", stats["nha_trang"] + stats["da_lat"])
    st.caption("Bảng này là số tổng của toàn công ty. Phần chi tiết từng bộ phận bên dưới phải cộng lại khớp với đây.")
    st.dataframe(company_frame, width="stretch", hide_index=True)
    st.bar_chart(company_frame.set_index("Hạng mục")[["Số lượng"]], width="stretch")
    st.markdown("</div>", unsafe_allow_html=True)

    department_frame = _department_summary_frame(st.session_state.records)
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Tổng hợp theo bộ phận</div>', unsafe_allow_html=True)
    if not department_frame.empty:
        st.caption("Mỗi dòng là tổng của một bộ phận. Dòng TỔNG bên dưới phải khớp với bảng tổng quan toàn công ty.")
        st.dataframe(department_frame, width="stretch", hide_index=True)
        chart_source = department_frame[department_frame["Bộ phận"] != "TỔNG"].set_index("Bộ phận")
        st.bar_chart(chart_source[["Tổng phiếu", "Tham gia", "Không tham gia"]], width="stretch")
    else:
        st.info("Chưa có dữ liệu để hiển thị theo bộ phận.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Biểu đồ chi tiết từng bộ phận</div>', unsafe_allow_html=True)
    if not department_frame.empty:
        st.caption("Biểu đồ này cho biết từng bộ phận có bao nhiêu người tham gia, không tham gia, chọn Nha Trang và chọn Đà Lạt.")
        st.dataframe(department_frame, width="stretch", hide_index=True)
        detail_chart = department_frame[department_frame["Bộ phận"] != "TỔNG"].set_index("Bộ phận")
        st.bar_chart(detail_chart[["Tham gia", "Không tham gia", "Nha Trang", "Đà Lạt"]], width="stretch")
    else:
        st.info("Chưa có dữ liệu chi tiết từng bộ phận.")
    st.markdown("</div>", unsafe_allow_html=True)


render_shell()
from __future__ import annotations

from io import BytesIO
import json
import os
from pathlib import Path
from datetime import datetime
from typing import Any

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

from excel_export import build_workbook

try:
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    from googleapiclient.http import MediaFileUpload
except ImportError:
    Credentials = None
    build = None
    MediaIoBaseDownload = None
    MediaFileUpload = None


APP_TITLE = "Khảo sát du lịch công ty Meiwa"
DESTINATIONS = ["Nha Trang", "Đà Lạt"]
DEPARTMENTS = ["GA", "CR", "CS", "CD", "PT", "QA", "MOLD"]
EXPORT_FILE_NAME = "Form khao sat Du lich Cong ty meiwa nam 2026.xlsx"
EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def init_state() -> None:
    if "records" not in st.session_state:
        st.session_state.records = []
    if "dirty_export" not in st.session_state:
        st.session_state.dirty_export = False
    if "last_saved_path" not in st.session_state:
        st.session_state.last_saved_path = ""
    if "last_saved_at" not in st.session_state:
        st.session_state.last_saved_at = ""
    if "last_drive_url" not in st.session_state:
        st.session_state.last_drive_url = ""
    if "input_msnv" not in st.session_state:
        st.session_state.input_msnv = ""
    if "input_ho_ten" not in st.session_state:
        st.session_state.input_ho_ten = ""
    if "input_bo_phan" not in st.session_state:
        st.session_state.input_bo_phan = DEPARTMENTS[3]
    if "input_cong_doan" not in st.session_state:
        st.session_state.input_cong_doan = ""
    if "input_tham_gia" not in st.session_state:
        st.session_state.input_tham_gia = None
    if "input_dia_diem" not in st.session_state:
        st.session_state.input_dia_diem = None
    if "reset_form_pending" not in st.session_state:
        st.session_state.reset_form_pending = False


def reset_form_inputs() -> None:
    st.session_state.input_msnv = ""
    st.session_state.input_ho_ten = ""
    st.session_state.input_bo_phan = DEPARTMENTS[3]
    st.session_state.input_cong_doan = ""
    st.session_state.input_tham_gia = None
    st.session_state.input_dia_diem = None


def normalize_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_name(value: object) -> str:
    raw = " ".join(normalize_text(value).split())
    if not raw:
        return ""
    return " ".join(part[:1].upper() + part[1:].lower() if part else "" for part in raw.split(" "))


def normalize_department(value: object) -> str:
    raw = normalize_text(value).upper().replace(".", "").replace(" ", "")
    if not raw:
        return ""
    aliases = {
        "MOLD": "MOLD",
        "GA": "GA",
        "CR": "CR",
        "CS": "CS",
        "CD": "CD",
        "PT": "PT",
        "QA": "QA",
    }
    return aliases.get(raw, raw)


def is_joining(value: object) -> bool:
    return normalize_text(value).lower() in {"co", "có", "yes", "y", "1", "true"}


def record_df(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(
            columns=["MSNV", "Họ tên", "Bộ phận", "Công đoạn", "Tham gia", "Địa điểm"]
        )
    frame = pd.DataFrame(records)
    frame = frame.rename(
        columns={
            "msnv": "MSNV",
            "ho_ten": "Họ tên",
            "bo_phan": "Bộ phận",
            "cong_doan": "Công đoạn",
            "tham_gia": "Tham gia",
            "dia_diem": "Địa điểm",
        }
    )
    wanted = ["MSNV", "Họ tên", "Bộ phận", "Công đoạn", "Tham gia", "Địa điểm"]
    for col in wanted:
        if col not in frame.columns:
            frame[col] = ""
    frame = frame[wanted]
    return frame


def upsert_record(record: dict) -> None:
    records = st.session_state.records
    key = normalize_text(record.get("msnv"))
    if not key:
        return
    replaced = False
    for idx, existing in enumerate(records):
        if normalize_text(existing.get("msnv")) == key:
            records[idx] = record
            replaced = True
            break
    if not replaced:
        records.append(record)


def metrics(records: list[dict]) -> dict[str, int]:
    yes = 0
    no = 0
    nha_trang = 0
    da_lat = 0
    for item in records:
        participate = normalize_text(item.get("tham_gia")).lower() in {"co", "có", "yes", "y", "1", "true"}
        destination = normalize_text(item.get("dia_diem")).lower()
        if participate:
            yes += 1
            if destination == "nha trang":
                nha_trang += 1
            elif destination in {"đà lạt", "da lat"}:
                da_lat += 1
        else:
            no += 1
    return {
        "total": len(records),
        "yes": yes,
        "no": no,
        "nha_trang": nha_trang,
        "da_lat": da_lat,
    }


def department_summary(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=["Bộ phận", "Tổng phiếu", "Tham gia"])

    frame = pd.DataFrame(records)
    if "bo_phan" in frame.columns:
        frame["bo_phan"] = frame["bo_phan"].map(normalize_department)
    else:
        frame["bo_phan"] = ""
    if "tham_gia" in frame.columns:
        frame["tham_gia"] = frame["tham_gia"].astype(str).str.strip().str.lower()
    else:
        frame["tham_gia"] = ""
    frame["is_join"] = frame["tham_gia"].isin(["co", "có", "yes", "y", "1", "true"])

    grouped = (
        frame.groupby("bo_phan", dropna=False)
        .agg(Tổng_phiếu=("msnv", "count"), Tham_gia=("is_join", "sum"))
        .reset_index()
        .rename(columns={"bo_phan": "Bộ phận", "Tổng_phiếu": "Tổng phiếu", "Tham_gia": "Tham gia"})
    )

    if not grouped.empty:
        order_map = {dept: idx for idx, dept in enumerate(DEPARTMENTS)}
        grouped["_order"] = grouped["Bộ phận"].map(lambda x: order_map.get(str(x).strip().upper(), 999))
        grouped = grouped.sort_values(["_order", "Bộ phận"]).drop(columns=["_order"]).reset_index(drop=True)
    return grouped


def destination_summary(records: list[dict]) -> pd.DataFrame:
    stats = metrics(records)
    total_join = stats["yes"]
    rows = [
        {
            "Hạng mục": "Tổng tham gia",
            "Số người": total_join,
            "Tỷ lệ": "100%" if total_join else "0%",
        },
        {
            "Hạng mục": "Nha Trang",
            "Số người": stats["nha_trang"],
            "Tỷ lệ": f"{(stats['nha_trang'] / total_join * 100):.1f}%" if total_join else "0%",
        },
        {
            "Hạng mục": "Đà Lạt",
            "Số người": stats["da_lat"],
            "Tỷ lệ": f"{(stats['da_lat'] / total_join * 100):.1f}%" if total_join else "0%",
        },
    ]
    return pd.DataFrame(rows)


def department_destination_summary(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=["Bộ phận", "Tham gia", "Nha Trang", "Đà Lạt"])

    rows: list[dict] = []
    for department in DEPARTMENTS:
        dept_records = [
            item for item in records if normalize_department(item.get("bo_phan")) == department
        ]
        stats = metrics(dept_records)
        if stats["total"] == 0:
            continue
        rows.append(
            {
                "Bộ phận": department,
                "Tham gia": stats["yes"],
                "Nha Trang": stats["nha_trang"],
                "Đà Lạt": stats["da_lat"],
            }
        )
    return pd.DataFrame(rows)


def import_rows_from_frame(frame: pd.DataFrame) -> list[dict]:
    if frame.empty:
        return []

    lookup = {str(col).strip().lower().replace(" ", "").replace("_", ""): col for col in frame.columns}

    def pick(*options: str):
        for option in options:
            key = option.strip().lower().replace(" ", "").replace("_", "")
            if key in lookup:
                return lookup[key]
        return None

    msnv_col = pick("msnv", "ma nv", "manv", "employee id")
    name_col = pick("họ tên", "ho ten", "hoten", "name")
    dept_col = pick("bộ phận", "bo phan", "bophan", "department")
    dept_process_col = pick("công đoạn", "cong doan", "process", "stage")
    join_col = pick("tham gia", "thamgia", "join", "participate")
    dest_col = pick("địa điểm", "dia diem", "destination")

    rows: list[dict] = []
    for _, row in frame.iterrows():
        msnv = normalize_text(row[msnv_col]) if msnv_col else ""
        name = normalize_name(row[name_col]) if name_col else ""
        if not (msnv or name):
            continue
        join_raw = normalize_text(row[join_col]) if join_col else ""
        join_raw_lower = join_raw.lower()
        if (not join_raw) and (normalize_text(row[dest_col]) if dest_col else ""):
            join_raw = "Có"
            join_raw_lower = "có"
        participate = join_raw_lower in {"co", "có", "yes", "y", "1", "true", "x", "v"}
        destination = normalize_text(row[dest_col]) if dest_col else ""
        if not participate:
            destination = ""
            people = 0
        rows.append(
            {
                "msnv": msnv,
                "ho_ten": normalize_name(name),
                "bo_phan": normalize_department(row[dept_col]) if dept_col else "",
                "cong_doan": normalize_text(row[dept_process_col]) if dept_process_col else "",
                "tham_gia": "Có" if participate else "Không",
                "dia_diem": destination if participate else "",
            }
        )
    return rows


def build_export_bytes(records: list[dict]) -> bytes:
    workbook = build_workbook(records)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def get_export_path() -> Path:
    data_dir = os.getenv("DATA_DIR", "")
    export_dir = Path(data_dir) if data_dir else Path(__file__).with_name("exports")
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir / EXPORT_FILE_NAME


def load_existing_records() -> list[dict]:
    export_path = get_export_path()
    if not export_path.exists():
        return []

    workbook = load_workbook(export_path, data_only=False)
    sheet = workbook.active
    records: list[dict] = []

    for row_num in range(9, sheet.max_row + 1):
        msnv = normalize_text(sheet.cell(row_num, 2).value)
        if msnv.upper() == "TOTAL":
            break

        ho_ten = normalize_name(sheet.cell(row_num, 3).value)
        bo_phan = normalize_department(sheet.cell(row_num, 4).value)
        cong_doan = normalize_text(sheet.cell(row_num, 5).value)
        di_co = normalize_text(sheet.cell(row_num, 6).value).lower() == "v"
        di_khong = normalize_text(sheet.cell(row_num, 7).value).lower() == "v"
        nha_trang = normalize_text(sheet.cell(row_num, 8).value).lower() == "v"
        da_lat = normalize_text(sheet.cell(row_num, 9).value).lower() == "v"

        if not any([msnv, ho_ten, bo_phan, cong_doan, di_co, di_khong, nha_trang, da_lat]):
            continue

        tham_gia = "Có" if di_co else "Không"
        dia_diem = "Nha Trang" if nha_trang else "Đà Lạt" if da_lat else ""
        records.append(
            {
                "msnv": msnv,
                "ho_ten": ho_ten,
                "bo_phan": bo_phan,
                "cong_doan": cong_doan,
                "tham_gia": tham_gia,
                "dia_diem": dia_diem if tham_gia == "Có" else "",
            }
        )

    return records


def cleanup_old_export_files(export_dir: Path, keep_name: str) -> None:
    for file_path in export_dir.glob("*.xlsx"):
        if file_path.name != keep_name:
            try:
                file_path.unlink()
            except OSError:
                continue


def read_google_config_from_env() -> tuple[dict[str, Any], dict[str, Any]]:
    drive_settings = {
        "folder_id": os.getenv("GOOGLE_DRIVE_FOLDER_ID", ""),
        "shared_url": os.getenv("GOOGLE_DRIVE_SHARED_URL", ""),
    }
    service_account_raw = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    service_account: dict[str, Any] = {}
    if service_account_raw:
        try:
            service_account = json.loads(service_account_raw)
        except json.JSONDecodeError:
            service_account = {}
    return drive_settings, service_account


def get_google_drive_settings() -> dict[str, str]:
    drive_settings: Any = {}
    service_account: Any = {}

    try:
        if "google_drive" in st.secrets:
            drive_settings = st.secrets["google_drive"]
        if "google_service_account" in st.secrets:
            service_account = st.secrets["google_service_account"]
    except Exception:
        drive_settings, service_account = read_google_config_from_env()
    else:
        if not drive_settings and not service_account:
            drive_settings, service_account = read_google_config_from_env()

    folder_id = normalize_text(drive_settings.get("folder_id", ""))
    shared_url = normalize_text(drive_settings.get("shared_url", ""))
    enabled = bool(folder_id and service_account)

    return {
        "folder_id": folder_id,
        "shared_url": shared_url,
        "enabled": "1" if enabled else "0",
    }


@st.cache_resource(show_spinner=False)
def get_google_drive_service():
    if Credentials is None or build is None:
        return None
    service_account_info: dict[str, Any] | None = None
    try:
        has_service_account = "google_service_account" in st.secrets
    except Exception:
        has_service_account = False
    if has_service_account:
        service_account_info = dict(st.secrets["google_service_account"])
    else:
        _, env_service_account = read_google_config_from_env()
        if env_service_account:
            service_account_info = env_service_account
    if not service_account_info:
        return None
    credentials = Credentials.from_service_account_info(service_account_info, scopes=["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=credentials)


def sync_export_to_google_drive(local_path: Path) -> dict[str, str]:
    settings = get_google_drive_settings()
    if settings["enabled"] != "1":
        return {"status": "not_configured", "url": settings["shared_url"]}

    service = get_google_drive_service()
    if service is None or MediaFileUpload is None:
        return {"status": "missing_library", "url": settings["shared_url"]}

    file_name = local_path.name.replace("'", "\\'")
    query = (
        f"name = '{file_name}' and "
        f"'{settings['folder_id']}' in parents and trashed = false"
    )
    existing = (
        service.files()
        .list(
            q=query,
            spaces="drive",
            fields="files(id, webViewLink, webContentLink)",
            pageSize=1,
        )
        .execute()
        .get("files", [])
    )
    media = MediaFileUpload(str(local_path), mimetype=EXCEL_MIME_TYPE, resumable=False)

    if existing:
        file_id = existing[0]["id"]
        updated = (
            service.files()
            .update(fileId=file_id, media_body=media, fields="id, webViewLink, webContentLink")
            .execute()
        )
        url = updated.get("webViewLink") or updated.get("webContentLink") or settings["shared_url"]
        return {"status": "updated", "url": url}

    created = (
        service.files()
        .create(
            body={"name": local_path.name, "parents": [settings["folder_id"]]},
            media_body=media,
            fields="id, webViewLink, webContentLink",
        )
        .execute()
    )
    url = created.get("webViewLink") or created.get("webContentLink") or settings["shared_url"]
    return {"status": "created", "url": url}


def ensure_local_export_from_google_drive() -> dict[str, str]:
    export_path = get_export_path()
    if export_path.exists():
        return {"status": "local_exists", "path": str(export_path)}

    settings = get_google_drive_settings()
    if settings["enabled"] != "1":
        return {"status": "not_configured", "path": str(export_path)}

    service = get_google_drive_service()
    if service is None or MediaIoBaseDownload is None:
        return {"status": "missing_library", "path": str(export_path)}

    file_name = export_path.name.replace("'", "\\'")
    query = (
        f"name = '{file_name}' and "
        f"'{settings['folder_id']}' in parents and trashed = false"
    )
    existing = (
        service.files()
        .list(
            q=query,
            spaces="drive",
            fields="files(id, name)",
            pageSize=1,
        )
        .execute()
        .get("files", [])
    )
    if not existing:
        return {"status": "drive_file_missing", "path": str(export_path)}

    request = service.files().get_media(fileId=existing[0]["id"])
    export_path.parent.mkdir(parents=True, exist_ok=True)
    with export_path.open("wb") as file_obj:
        downloader = MediaIoBaseDownload(file_obj, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    return {"status": "downloaded", "path": str(export_path)}


def save_export_file(records: list[dict]) -> Path:
    export_dir = get_export_path().parent
    cleanup_old_export_files(export_dir, EXPORT_FILE_NAME)
    export_path = get_export_path()
    workbook = build_workbook(records)
    workbook.save(export_path)
    return export_path


def render_close_warning(is_dirty: bool) -> None:
    components.html(
        f"""
        <script>
        window.onbeforeunload = {str(is_dirty).lower()}
          ? function(event) {{
              event.preventDefault();
              event.returnValue = "";
              return "";
            }}
          : null;
        </script>
        """,
        height=0,
    )


def load_rows_from_upload(uploaded_file) -> list[dict]:
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix == ".csv":
        frame = pd.read_csv(uploaded_file)
    else:
        frame = pd.read_excel(uploaded_file)
    return import_rows_from_frame(frame)


def department_summary(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=["B盻・ph蘯ｭn", "T盻貧g phi蘯ｿu", "Tham gia", "Khﾃｴng tham gia"])

    frame = pd.DataFrame(records)
    frame["bo_phan"] = frame.get("bo_phan", "").map(normalize_department) if "bo_phan" in frame.columns else ""
    frame["tham_gia"] = frame.get("tham_gia", "").astype(str).str.strip().str.lower() if "tham_gia" in frame.columns else ""
    frame["is_join"] = frame["tham_gia"].isin(["co", "cﾃｳ", "yes", "y", "1", "true"])

    grouped = (
        frame.groupby("bo_phan", dropna=False)
        .agg(T盻貧g_phi蘯ｿu=("msnv", "count"), Tham_gia=("is_join", "sum"))
        .reset_index()
        .rename(columns={"bo_phan": "B盻・ph蘯ｭn", "T盻貧g_phi蘯ｿu": "T盻貧g phi蘯ｿu", "Tham_gia": "Tham gia"})
    )
    grouped["Khﾃｴng tham gia"] = grouped["T盻貧g phi蘯ｿu"] - grouped["Tham gia"]

    order_map = {dept: idx for idx, dept in enumerate(DEPARTMENTS)}
    grouped["_order"] = grouped["B盻・ph蘯ｭn"].map(lambda x: order_map.get(str(x).strip().upper(), 999))
    return grouped.sort_values(["_order", "B盻・ph蘯ｭn"]).drop(columns=["_order"]).reset_index(drop=True)


def destination_summary(records: list[dict]) -> pd.DataFrame:
    stats = metrics(records)
    total_join = stats["yes"]
    rows = [
        {
            "H蘯｡ng m盻･c": "T盻貧g tham gia",
            "S盻・ngﾆｰ盻拱": total_join,
            "T盻ｷ l盻・": "100%" if total_join else "0%",
        },
        {
            "H蘯｡ng m盻･c": "Khﾃｴng tham gia",
            "S盻・ngﾆｰ盻拱": stats["no"],
            "T盻ｷ l盻・": f"{(stats['no'] / stats['total'] * 100):.1f}%" if stats["total"] else "0%",
        },
        {
            "H蘯｡ng m盻･c": "Nha Trang",
            "S盻・ngﾆｰ盻拱": stats["nha_trang"],
            "T盻ｷ l盻・": f"{(stats['nha_trang'] / total_join * 100):.1f}%" if total_join else "0%",
        },
        {
            "H蘯｡ng m盻･c": "ﾄ静 L蘯｡t",
            "S盻・ngﾆｰ盻拱": stats["da_lat"],
            "T盻ｷ l盻・": f"{(stats['da_lat'] / total_join * 100):.1f}%" if total_join else "0%",
        },
    ]
    return pd.DataFrame(rows)


def department_destination_summary(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=["B盻・ph蘯ｭn", "Tham gia", "Khﾃｴng tham gia", "Nha Trang", "ﾄ静 L蘯｡t"])

    rows: list[dict] = []
    for department in DEPARTMENTS:
        dept_records = [
            item for item in records if normalize_department(item.get("bo_phan")) == department
        ]
        stats = metrics(dept_records)
        if stats["total"] == 0:
            continue
        rows.append(
            {
                "B盻・ph蘯ｭn": department,
                "Tham gia": stats["yes"],
                "Khﾃｴng tham gia": stats["no"],
                "Nha Trang": stats["nha_trang"],
                "ﾄ静 L蘯｡t": stats["da_lat"],
            }
        )
    return pd.DataFrame(rows)


def render_shell() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="🧳", layout="centered")
    st.markdown(
        """
        <style>
        .block-container {
            padding-top: 1.25rem;
            padding-bottom: 2rem;
            max-width: 760px;
        }
        .hero {
            border-radius: 24px;
            padding: 1.2rem 1.4rem;
            background: linear-gradient(135deg, #fff6ea 0%, #eef6ff 48%, #f6ecff 100%);
            border: 1px solid rgba(17, 24, 39, 0.08);
            box-shadow: 0 16px 40px rgba(15, 23, 42, 0.08);
            margin-bottom: 1rem;
        }
        .hero h1 {
            margin: 0;
            font-size: 2rem;
            line-height: 1.15;
        }
        .hero p {
            margin: 0.35rem 0 0;
            color: #475569;
        }
        .card {
            background: #ffffff;
            border: 1px solid rgba(15, 23, 42, 0.08);
            border-radius: 20px;
            padding: 1rem 1rem 0.85rem;
            box-shadow: 0 10px 28px rgba(15, 23, 42, 0.06);
            margin-bottom: 1rem;
        }
        .save-alert {
            border: 2px solid #dc2626;
            background: linear-gradient(135deg, #fff7ed 0%, #fef2f2 100%);
        }
        .save-title {
            color: #b91c1c;
            font-size: 1.1rem;
            font-weight: 800;
            margin-bottom: 0.35rem;
        }
        .save-copy {
            color: #7f1d1d;
            font-size: 0.95rem;
            margin-bottom: 0.8rem;
        }
        .section-title {
            font-size: 1rem;
            font-weight: 700;
            margin: 0 0 0.6rem;
        }
        .hint {
            color: #64748b;
            font-size: 0.92rem;
            margin-top: 0.25rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    init_state()
    if st.session_state.reset_form_pending:
        reset_form_inputs()
        st.session_state.reset_form_pending = False
    if "bootstrap_drive_status" not in st.session_state:
        st.session_state.bootstrap_drive_status = ensure_local_export_from_google_drive().get("status", "")
    st.session_state.records = load_existing_records()
    render_close_warning(st.session_state.dirty_export)

    st.markdown(
        f"""
        <div class="hero">
          <h1>{APP_TITLE}</h1>
          <p>Mở app là nhập ngay. Nếu chọn Không, phần địa điểm tự ẩn.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="card">', unsafe_allow_html=True)
    msnv = st.text_input("MSNV", placeholder="Ví dụ: G06071209", key="input_msnv")
    ho_ten = st.text_input("Họ tên", placeholder="Ví dụ: Lê Hữu Phước", key="input_ho_ten")
    bo_phan = st.selectbox("Bộ phận", DEPARTMENTS, key="input_bo_phan")
    cong_doan = st.text_input("Công đoạn", placeholder="Ví dụ: 000 hoặc 3-5", key="input_cong_doan")
    tham_gia = st.selectbox(
        "Bạn có tham gia chuyến đi không?",
        ["Có", "Không"],
        index=None,
        placeholder="Chọn Có hoặc Không",
        key="input_tham_gia",
    )
    dia_diem = ""
    if is_joining(tham_gia):
        dia_diem = st.selectbox(
            "Chọn 1 địa điểm",
            DESTINATIONS,
            index=None,
            placeholder="Chọn Nha Trang hoặc Đà Lạt",
            key="input_dia_diem",
        )

    submitted = st.button("LƯU KẾT QUẢ KHẢO SÁT", width="stretch")
    if submitted:
        if not msnv or not ho_ten:
            st.error("Nhập ít nhất MSNV và Họ tên.")
        elif not tham_gia:
            st.error("Chọn Có hoặc Không tham gia.")
        elif is_joining(tham_gia) and not dia_diem:
            st.error("Chọn địa điểm khi người đó tham gia.")
        else:
            record = {
                "msnv": msnv,
                "ho_ten": normalize_name(ho_ten),
                "bo_phan": normalize_department(bo_phan),
                "cong_doan": cong_doan,
                "tham_gia": "Có" if is_joining(tham_gia) else "Không",
                "dia_diem": dia_diem if is_joining(tham_gia) else "",
            }
            st.session_state.records = load_existing_records()
            upsert_record(record)
            export_path = save_export_file(st.session_state.records)
            drive_result = {"status": "not_configured", "url": ""}
            try:
                drive_result = sync_export_to_google_drive(export_path)
            except Exception as exc:
                drive_result = {"status": f"error: {exc}", "url": ""}
            st.session_state.last_saved_path = str(export_path)
            st.session_state.last_saved_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            st.session_state.last_drive_url = drive_result.get("url", "")
            st.session_state.dirty_export = False
            st.session_state.reset_form_pending = True
            if drive_result["status"] in {"created", "updated"}:
                st.success(
                    f"Đã lưu khảo sát, cập nhật file tổng và đồng bộ Google Drive lúc {st.session_state.last_saved_at}"
                )
            elif drive_result["status"] == "not_configured":
                st.success(
                    f"Đã lưu khảo sát và cập nhật file tổng lúc {st.session_state.last_saved_at}"
                )
            elif drive_result["status"] == "missing_library":
                st.warning("Đã lưu file tổng, nhưng máy chủ chưa cài thư viện Google Drive để đồng bộ.")
            else:
                st.warning(
                    f"Đã lưu file tổng, nhưng chưa đẩy được lên Google Drive: {drive_result['status']}"
                )
            st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card save-alert">', unsafe_allow_html=True)
    st.markdown('<div class="save-title">TỰ ĐỘNG LƯU FILE TỔNG</div>', unsafe_allow_html=True)
    st.markdown('<div class="save-copy">Mỗi lần bấm "LƯU KẾT QUẢ KHẢO SÁT", app sẽ tự cập nhật ngay vào file Excel tổng.</div>', unsafe_allow_html=True)
    excel_bytes = build_export_bytes(st.session_state.records)
    st.download_button(
        "TẢI FILE TỔNG",
        data=excel_bytes,
        file_name=EXPORT_FILE_NAME,
        mime=EXCEL_MIME_TYPE,
        width="stretch",
    )
    if st.session_state.last_saved_path:
        st.caption(f"File tổng mới nhất: `{st.session_state.last_saved_path}`")
        st.markdown(f"[Mở file tổng đã lưu]({Path(st.session_state.last_saved_path).as_uri()})")
    drive_settings = get_google_drive_settings()
    if drive_settings["enabled"] == "1":
        st.caption("Google Drive: đã cấu hình đồng bộ tự động.")
    else:
        st.caption("Google Drive: chưa cấu hình. App hiện vẫn lưu file tổng trong máy và cho tải trực tiếp.")
    drive_url = st.session_state.last_drive_url or drive_settings["shared_url"]
    if drive_url:
        st.markdown(f"[MỞ FILE TỔNG TRÊN GOOGLE DRIVE]({drive_url})")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Đã nhập</div>', unsafe_allow_html=True)
    if st.session_state.records:
        rows = [
            {
                "MSNV": item.get("msnv", ""),
                "Họ tên": item.get("ho_ten", ""),
                "Bộ phận": item.get("bo_phan", ""),
                "Công đoạn": item.get("cong_doan", ""),
                "Đi": item.get("tham_gia", ""),
                "Địa điểm": item.get("dia_diem", ""),
            }
            for item in st.session_state.records
        ]
        st.dataframe(rows, width="stretch", hide_index=True)
    else:
        st.info("Chưa có dữ liệu.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Biểu đồ tổng toàn công ty</div>', unsafe_allow_html=True)
    stats = metrics(st.session_state.records)
    overview = destination_summary(st.session_state.records)
    if stats["yes"] > 0:
        col1, col2, col3 = st.columns(3)
        col1.metric("Tổng người tham gia", stats["yes"])
        col2.metric("Nha Trang", stats["nha_trang"], f"{(stats['nha_trang'] / stats['yes'] * 100):.1f}%")
        col3.metric("Đà Lạt", stats["da_lat"], f"{(stats['da_lat'] / stats['yes'] * 100):.1f}%")
        st.dataframe(overview, width="stretch", hide_index=True)
        chart_frame = overview[overview["Hạng mục"] != "Tổng tham gia"].set_index("Hạng mục")[["Số người"]]
        st.bar_chart(chart_frame, width="stretch")
    else:
        st.info("Chưa có người chọn tham gia để hiển thị biểu đồ địa điểm.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Biểu đồ từng bộ phận</div>', unsafe_allow_html=True)
    dept_frame = department_summary(st.session_state.records)
    if not dept_frame.empty:
        st.dataframe(dept_frame, width="stretch", hide_index=True)
        chart_frame = dept_frame.set_index("Bộ phận")[["Tham gia"]]
        st.bar_chart(chart_frame, width="stretch")
        destination_by_department = department_destination_summary(st.session_state.records)
        if not destination_by_department.empty:
            st.dataframe(destination_by_department, width="stretch", hide_index=True)
            detail_chart_frame = destination_by_department.set_index("Bộ phận")[["Nha Trang", "Đà Lạt"]]
            st.bar_chart(detail_chart_frame, width="stretch")
    else:
        st.info("Chưa có dữ liệu để hiển thị dashboard bộ phận.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Bi盻ブ b盻・sung: tham gia vﾃ khﾃｴng tham gia</div>', unsafe_allow_html=True)
    company_compare = pd.DataFrame(
        [
            {"H蘯｡ng m盻･c": "Tham gia", "S盻・ngﾆｰ盻拱": stats["yes"]},
            {"H蘯｡ng m盻･c": "Khﾃｴng tham gia", "S盻・ngﾆｰ盻拱": stats["no"]},
            {"H蘯｡ng m盻･c": "Nha Trang", "S盻・ngﾆｰ盻拱": stats["nha_trang"]},
            {"H蘯｡ng m盻･c": "ﾄ静 L蘯｡t", "S盻・ngﾆｰ盻拱": stats["da_lat"]},
        ]
    )
    st.dataframe(company_compare, width="stretch", hide_index=True)
    st.bar_chart(company_compare.set_index("H蘯｡ng m盻･c")[["S盻・ngﾆｰ盻拱"]], width="stretch")

    destination_by_department = department_destination_summary(st.session_state.records)
    if not destination_by_department.empty:
        st.markdown('<div class="hint">Chi ti盻ｿt theo b盻・ph蘯ｭn: hi盻ハ th盻・c蘯｣ tham gia, khﾃｴng tham gia vﾃ ﾄ黛ｻ蟻 ﾄ訴盻ノ ﾄ妥｣ ch盻肱.</div>', unsafe_allow_html=True)
        st.dataframe(destination_by_department, width="stretch", hide_index=True)
        st.bar_chart(
            destination_by_department.set_index("B盻・ph蘯ｭn")[["Tham gia", "Khﾃｴng tham gia", "Nha Trang", "ﾄ静 L蘯｡t"]],
            width="stretch",
        )
    st.markdown("</div>", unsafe_allow_html=True)


    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Dashboard tong hop de doc</div>', unsafe_allow_html=True)
    clean_company = pd.DataFrame(
        [
            {"Hang muc": "Tong phieu", "So nguoi": stats["total"]},
            {"Hang muc": "Tong tham gia", "So nguoi": stats["yes"]},
            {"Hang muc": "Khong tham gia", "So nguoi": stats["no"]},
            {"Hang muc": "Nha Trang", "So nguoi": stats["nha_trang"]},
            {"Hang muc": "Da Lat", "So nguoi": stats["da_lat"]},
        ]
    )
    st.caption("Bang tong hop toan cong ty de doi chieu nhanh.")
    st.dataframe(clean_company, width="stretch", hide_index=True)
    st.bar_chart(clean_company.set_index("Hang muc")[["So nguoi"]], width="stretch")

    clean_department_rows = []
    for dept in DEPARTMENTS:
        dept_items = [
            item for item in st.session_state.records if normalize_department(item.get("bo_phan")) == dept
        ]
        dept_stats = metrics(dept_items)
        if dept_stats["total"] == 0:
            continue
        clean_department_rows.append(
            {
                "Bo phan": dept,
                "Tong phieu": dept_stats["total"],
                "Tham gia": dept_stats["yes"],
                "Khong tham gia": dept_stats["no"],
                "Nha Trang": dept_stats["nha_trang"],
                "Da Lat": dept_stats["da_lat"],
            }
        )

    clean_department = pd.DataFrame(clean_department_rows)
    if not clean_department.empty:
        total_detail = pd.DataFrame(
            [
                {
                    "Bo phan": "TONG",
                    "Tong phieu": int(clean_department["Tong phieu"].sum()),
                    "Tham gia": int(clean_department["Tham gia"].sum()),
                    "Khong tham gia": int(clean_department["Khong tham gia"].sum()),
                    "Nha Trang": int(clean_department["Nha Trang"].sum()),
                    "Da Lat": int(clean_department["Da Lat"].sum()),
                }
            ]
        )
        clean_department_view = pd.concat([clean_department, total_detail], ignore_index=True)
        st.markdown('<div class="section-title">Chi tiet tung bo phan co dong TONG</div>', unsafe_allow_html=True)
        st.caption("Dong TONG ben duoi phai khop voi so tong quat ben tren.")
        st.dataframe(clean_department_view, width="stretch", hide_index=True)
        st.bar_chart(
            clean_department.set_index("Bo phan")[["Tong phieu", "Tham gia", "Khong tham gia", "Nha Trang", "Da Lat"]],
            width="stretch",
        )
    st.markdown("</div>", unsafe_allow_html=True)


render_shell()
