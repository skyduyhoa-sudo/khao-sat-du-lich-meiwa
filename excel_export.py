from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill


TEMPLATE_PATH = Path(__file__).with_name("Form khao sat Du lich Cong ty meiwa nam 2026.xlsx")


def _copy_cell_style(source, target) -> None:
    target._style = copy(source._style)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.number_format:
        target.number_format = source.number_format
    if source.protection:
        target.protection = copy(source.protection)


def _copy_dimensions(source_ws, target_ws) -> None:
    for key, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].width = dim.width
        target_ws.column_dimensions[key].hidden = dim.hidden

    for row_idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = dim.height
        target_ws.row_dimensions[row_idx].hidden = dim.hidden


def _copy_static_area(source_ws, target_ws) -> None:
    static_ranges = [
        "A1:I4",
        "A6:I8",
    ]
    for cell_range in static_ranges:
        for row in source_ws[cell_range]:
            for source_cell in row:
                target_cell = target_ws[source_cell.coordinate]
                target_cell.value = source_cell.value
                _copy_cell_style(source_cell, target_cell)


def _apply_base_merges(sheet) -> None:
    for merge in [
        "A3:I3",
        "A4:I4",
        "A7:A8",
        "B7:B8",
        "C7:C8",
        "D7:E8",
        "F7:G7",
        "H7:I7",
    ]:
        sheet.merge_cells(merge)


def _apply_summary_merges(sheet, summary_row: int) -> None:
    sheet.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row + 1, end_column=5)


def _style_data_row(target_ws, row_num: int, source_ws, row_style_source: int, no_participation: bool) -> None:
    for col in range(1, 10):
        source_cell = source_ws.cell(row_style_source, col)
        target_cell = target_ws.cell(row_num, col)
        _copy_cell_style(source_cell, target_cell)

    if no_participation:
        yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
        for col in (8, 9):
            target_ws.cell(row_num, col).fill = yellow
    target_ws.row_dimensions[row_num].height = source_ws.row_dimensions[row_style_source].height or 21


def build_workbook(records: Iterable[dict]) -> Workbook:
    records = list(records)
    template = load_workbook(TEMPLATE_PATH)
    template_ws = template.active

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    _copy_dimensions(template_ws, sheet)
    _copy_static_area(template_ws, sheet)
    _apply_base_merges(sheet)
    sheet.freeze_panes = template_ws.freeze_panes
    sheet.sheet_view.showGridLines = False

    # Preserve visible title styling from the template but leave the survey note editable.
    sheet["A3"] = template_ws["A3"].value
    sheet["A4"] = template_ws["A4"].value
    sheet["F6"] = template_ws["F6"].value

    data_start_row = 9
    template_capacity = 7
    last_data_row = max(data_start_row + len(records) - 1, data_start_row + template_capacity - 1)
    summary_row = last_data_row + 1
    pct_row = last_data_row + 2

    # Copy the data-row shell for the visible area.
    for row_num in range(data_start_row, last_data_row + 1):
        no_participation = False
        _style_data_row(sheet, row_num, template_ws, 10, no_participation)

    # Fill responses.
    for index, record in enumerate(records, start=0):
        row_num = data_start_row + index
        sheet.cell(row_num, 1).value = index + 1
        sheet.cell(row_num, 2).value = record.get("msnv", "")
        sheet.cell(row_num, 3).value = record.get("ho_ten", "")
        sheet.cell(row_num, 4).value = record.get("bo_phan", "")
        sheet.cell(row_num, 5).value = record.get("cong_doan", "")

        participate = str(record.get("tham_gia", "")).strip().lower() in {"co", "có", "yes", "y", "1", "true"}
        destination = str(record.get("dia_diem", "")).strip()

        if participate:
            sheet.cell(row_num, 6).value = "v"
            sheet.cell(row_num, 7).value = None
            sheet.cell(row_num, 8).value = "v" if destination.lower() == "nha trang" else None
            sheet.cell(row_num, 9).value = "v" if destination.lower() == "đà lạt" or destination.lower() == "da lat" else None
        else:
            sheet.cell(row_num, 6).value = None
            sheet.cell(row_num, 7).value = "v"
            sheet.cell(row_num, 8).value = None
            sheet.cell(row_num, 9).value = None
            yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
            sheet.cell(row_num, 8).fill = yellow
            sheet.cell(row_num, 9).fill = yellow

    # Make the final visible data row feel like the original bottom border.
    _style_data_row(sheet, last_data_row, template_ws, 15, False)
    if records:
        last_record = records[-1]
        participate = str(last_record.get("tham_gia", "")).strip().lower() in {"co", "có", "yes", "y", "1", "true"}
        if not participate:
            yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
            sheet.cell(last_data_row, 8).fill = yellow
            sheet.cell(last_data_row, 9).fill = yellow

    yes_count = 0
    no_count = 0
    nha_trang_count = 0
    da_lat_count = 0
    for record in records:
        participate = str(record.get("tham_gia", "")).strip().lower() in {"co", "có", "yes", "y", "1", "true"}
        destination = str(record.get("dia_diem", "")).strip().lower()
        if participate:
            yes_count += 1
            if destination == "nha trang":
                nha_trang_count += 1
            elif destination in {"đà lạt", "da lat"}:
                da_lat_count += 1
        else:
            no_count += 1

    total_records = len(records)
    participation_ratio = (yes_count / total_records) if total_records else 0
    no_ratio = (no_count / total_records) if total_records else 0
    nha_trang_ratio = (nha_trang_count / yes_count) if yes_count else 0
    da_lat_ratio = (da_lat_count / yes_count) if yes_count else 0

    # Summary block.
    _apply_summary_merges(sheet, summary_row)
    _copy_cell_style(template_ws["B16"], sheet[f"B{summary_row}"])
    _copy_cell_style(template_ws["F16"], sheet[f"F{summary_row}"])
    _copy_cell_style(template_ws["G16"], sheet[f"G{summary_row}"])
    _copy_cell_style(template_ws["H16"], sheet[f"H{summary_row}"])
    _copy_cell_style(template_ws["I16"], sheet[f"I{summary_row}"])

    _copy_cell_style(template_ws["B17"], sheet[f"B{pct_row}"])
    _copy_cell_style(template_ws["F17"], sheet[f"F{pct_row}"])
    _copy_cell_style(template_ws["G17"], sheet[f"G{pct_row}"])
    _copy_cell_style(template_ws["H17"], sheet[f"H{pct_row}"])
    _copy_cell_style(template_ws["I17"], sheet[f"I{pct_row}"])

    sheet[f"B{summary_row}"] = "TOTAL"
    sheet[f"F{summary_row}"] = yes_count
    sheet[f"G{summary_row}"] = no_count
    sheet[f"H{summary_row}"] = nha_trang_count
    sheet[f"I{summary_row}"] = da_lat_count

    sheet[f"F{pct_row}"] = participation_ratio
    sheet[f"G{pct_row}"] = no_ratio
    sheet[f"H{pct_row}"] = nha_trang_ratio
    sheet[f"I{pct_row}"] = da_lat_ratio

    for cell in [sheet[f"F{pct_row}"], sheet[f"G{pct_row}"], sheet[f"H{pct_row}"], sheet[f"I{pct_row}"]]:
        cell.number_format = "0.0%"

    # Keep important labels centered and readable.
    sheet[f"B{summary_row}"].alignment = Alignment(horizontal="center", vertical="center")

    return workbook
